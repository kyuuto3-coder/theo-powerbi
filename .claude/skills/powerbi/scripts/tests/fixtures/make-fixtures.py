#!/usr/bin/env python3
"""Generates test fixtures. Run once on the dev machine; outputs are committed."""
import csv, datetime, os, random
from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))
random.seed(7)

# --- apps.xlsx: mimics rawdata/App Data.xlsx (unnamed index col A, header row 1, dictionary sheet at B2)
wb = Workbook()
ws = wb.active; ws.title = "Data"
ws.append([None, "id", "track_name", "size_bytes", "price", "rating_count_tot", "user_rating", "prime_genre", "release_date"])
genres = ["Games", "Education", "Weather", "Productivity", "Music"]
for i in range(40):
    ws.append([i, 281000000 + i, f"App {i}", random.randint(10_000_000, 900_000_000),
               random.choice([0, 0.99, 2.99, 4.99]), random.randint(0, 200000),
               random.choice([3, 3.5, 4, 4.5, 5]), random.choice(genres),
               datetime.date(2016, 1, 1) + datetime.timedelta(days=random.randint(0, 700))])
for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
    row[0].number_format = "yyyy-mm-dd"
wd = wb.create_sheet("Data Defintion")
wd["B2"] = "Name"; wd["C2"] = "Description"
for r, (n, d) in enumerate([("id", "App ID"), ("track_name", "Track Name"), ("size_bytes", "App Size in Bytes"),
                            ("price", "Price"), ("rating_count_tot", "Number of Rating"), ("user_rating", "User Rating"),
                            ("prime_genre", "Prime Genre"), ("release_date", "Release Date")], start=3):
    wd.cell(row=r, column=2, value=n); wd.cell(row=r, column=3, value=d)
wb.save(os.path.join(HERE, "apps.xlsx"))

# --- sales_cp949.csv + regions.csv: a tiny star schema with Korean headers in CP949
with open(os.path.join(HERE, "regions.csv"), "w", encoding="utf-8", newline="") as f:
    w = csv.writer(f); w.writerow(["지역코드", "지역명", "권역"])
    for code, name, zone in [(1, "서울", "수도권"), (2, "부산", "영남"), (3, "대구", "영남"), (4, "광주", "호남")]:
        w.writerow([code, name, zone])
with open(os.path.join(HERE, "sales_cp949.csv"), "w", encoding="cp949", newline="") as f:
    w = csv.writer(f); w.writerow(["주문일", "지역코드", "제품", "수량", "매출액"])
    for i in range(30):
        d = datetime.date(2025, 1, 1) + datetime.timedelta(days=i * 7)
        w.writerow([d.isoformat(), random.randint(1, 4), random.choice(["A", "B", "C"]), random.randint(1, 20), round(random.uniform(1000, 90000), 2)])
print("fixtures written to", HERE)
